import readcsv as course_dictionary
from scheduler import course_scheduler
from collections import namedtuple
import unittest
import re
from collections import namedtuple
from openpyxl import load_workbook
import warnings

class TestCourseScheduler(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.course_dict = course_dictionary.create_course_dict()

        # a dictionary that includes an additional major: spanish
        cls.span_dict = add_span_major()

    def test_dictionary(self):
        # CODE PROVIDED BY DR. FISHER
        # Test to see if all prereqs are in the file.
        prereq_list = [single_course for vals in self.course_dict.values()
                       for some_prereqs in vals.prereqs for single_course in some_prereqs]
        for prereq in prereq_list:
            if prereq not in self.course_dict:
                print(prereq)
        for key in self.course_dict:
            #Test to see if every course has a term and credits.
            if not self.course_dict[key].terms or not self.course_dict[key].credits:
                print(key)
            #Test to see if a course's prereqs include the course itself
            if key in [course for prereq in self.course_dict[key].prereqs for course in prereq]:
                print(key);
        # Prints all the CS courses.
        # for key in self.course_dict:
        #     if key.program == 'CS':
        #         print('type:', type(key), self.course_dict[key])
        # Prints the entire dictionary.
        # course_dictionary.print_dict(self.course_dict)
        # print(self.course_dict[('CS', 'open3')])

    def test_impossible_goal(self):
        # the scheduler should return an empty conjunction if conditions of the goal cannot be satisfied in four years
        # there are 100 total ECON and EDUC courses, so it is impossible to take all in 4 years
        all_econ_educ = [(key.program, key.designation) for key in self.course_dict
                         if (key.program == 'ECON' or key.program == 'EDUC')]
        plan = course_scheduler(self.course_dict, all_econ_educ, [])
        self.assertEqual(plan, ())

    def test_no_goal(self):
        plan = course_scheduler(self.course_dict, [], [])
        self.assertEqual(plan, ())

    def test_proper_terms(self):
        # ensuring the correct years and terms are included
        # this goal should result in a 5-semester plan
        plan = course_scheduler(self.course_dict, [('MATH', '4110')], [])
        split_plan = split_by_term(plan)
        years = list(split_plan.keys())
        self.assertNotEqual(split_plan['Frosh']['Fall']['courses'], [])
        self.assertNotEqual(split_plan['Frosh']['Spring']['courses'], [])
        self.assertNotEqual(split_plan['Soph']['Fall']['courses'], [])
        self.assertNotEqual(split_plan['Soph']['Spring']['courses'], [])
        self.assertNotEqual(split_plan['Junior']['Fall']['courses'], [])
        self.assertEqual(split_plan['Junior']['Spring']['courses'], [])
        self.assertEqual(split_plan['Senior']['Fall']['courses'], [])
        self.assertEqual(split_plan['Senior']['Spring']['courses'], [])

    def test_goal_satisfied(self):
        # testing the scheduler when the provided goal has already been satisfied
        plan = course_scheduler(self.course_dict, [('CS','1101')], [('CS','1101')])
        self.assertEqual(plan, ())

    def test_one_class_goal(self):
        # sum of credits must not be less than 12 per term
        plan = course_scheduler(self.course_dict, [('CS', '3250')], [])
        split_plan = split_by_term(plan)
        # ensure proper number of credits each semester
        for year in split_plan:
            for term in split_plan[year]:
                self.assertTrue((12 <= split_plan[year][term]['credits'] <= 18) or (split_plan[year][term]['credits'] == 0))

    def test_initial_state(self):
        plan = course_scheduler(self.course_dict, [('CS', '1101'), ('SPAN', '1102'), ('SPAN', '3325')],
                                [('SPAN', '1101')])
        # the prereq for SPAN1102 is already satisfied so neither of its prereqs should be in the plan
        for course in plan:
            self.assertNotEqual(course[0], ('SPAN', '1101'))
            self.assertNotEqual(course[0], ('SPAN', '1100'))

    def test_simple_plan(self):
        # in this case, there is no ambiguity in the optimal terms to schedule the goal and its prereqs in
        plan = course_scheduler(self.course_dict, [('MATH', '2410')], [])
        split_plan = split_by_term(plan)
        self.assertTrue(('MATH', '1200') in split_plan['Frosh']['Fall']['courses']
                        or ('MATH', '1300') in split_plan['Frosh']['Fall']['courses'])
        if ('MATH', '1200') in split_plan['Frosh']['Fall']['courses']:
            self.assertTrue(('MATH', '1201') in split_plan['Frosh']['Spring']['courses'])
            self.assertTrue(('MATH', '2200') in split_plan['Soph']['Fall']['courses'])
            self.assertTrue(('MATH', '2300') in split_plan['Soph']['Spring']['courses'])
            self.assertTrue(('MATH', '2410') in split_plan['Junior']['Fall']['courses'])
        if ('MATH', '1300') in split_plan['Frosh']['Fall']['courses']:
            self.assertTrue(('MATH', '1301') in split_plan['Frosh']['Spring']['courses'])
            self.assertTrue(('MATH', '2300') in split_plan['Soph']['Fall']['courses'])
            self.assertTrue(('MATH', '2410') in split_plan['Soph']['Spring']['courses'])

    def test_logistics(self):
        # ensure that all scheduled courses are actual courses (ie. are not higher level requirements or empty)
        # and are offered during the semester they are scheduled
        plan = course_scheduler(self.course_dict, [('CS', 'major')], [])
        split_plan = split_by_term(plan)
        for year in split_plan:
            for term in split_plan[year]:
                for course in split_plan[year][term]['courses']:
                    terms = self.course_dict[course.name].terms
                    self.assertTrue(term in terms)
                    self.assertNotEqual(self.course_dict[course.name].credits, 0)
                credit = split_plan[year][term]['credits']
                self.assertTrue((12 <= credit <= 18) or credit == 0)

    def test_5_credits(self):
        # goal of only 4 courses, but cannot all fit in one term since they're all 5-credit
        plan = course_scheduler(self.course_dict,
                                 [('SPAN', '1100'), ('SPAN', '1101'), ('SPAN', '1103'), ('SPAN', '2203')], [])
        split_plan = split_by_term(plan)
        self.assertNotEqual(split_plan['Frosh']['Fall']['courses'], [])
        self.assertNotEqual(split_plan['Frosh']['Spring']['courses'], [])
        self.assertEqual(split_plan['Soph']['Fall']['courses'], [])
        self.assertEqual(split_plan['Soph']['Spring']['courses'], [])
        self.assertEqual(split_plan['Junior']['Fall']['courses'], [])
        self.assertEqual(split_plan['Junior']['Spring']['courses'], [])
        self.assertEqual(split_plan['Senior']['Fall']['courses'], [])
        self.assertEqual(split_plan['Senior']['Spring']['courses'], [])

    def test_open_elective(self):
        # testing the open elective issue
        plan = course_scheduler(self.course_dict, [('CS', 'openelectives')], [])
        unique_courses = set()
        for course in plan:
            unique_courses.add(course[0].name)
        self.assertEqual(len(unique_courses), len(plan))

    """
    def test_spanish_major(self):
        # more thoroughly testing the spanish major plan (according to the catalog and project spec it is very simple)
        plan = course_scheduler(self.span_dict, [('SPAN', 'major')], [])
        split_plan = split_by_term(plan)
        # ensure proper number of credits each semester
        for year in split_plan:
            for term in split_plan[year]:
                self.assertTrue(12 <= split_plan[year][term]['credits'] <= 18)
        # check that all core requirements are satisfied
        self.assertTrue((('SPAN', '3301W'), ('Fall', 'Frosh'), 3) in plan)
        self.assertTrue((('SPAN', '3302'), ('Spring', 'Frosh'), 3) in plan)
        self.assertTrue((('SPAN', '3303'), ('Fall', 'Soph'), 3) in plan)
        Course = namedtuple('Course', 'program, designation')
        literature_req = self.span_dict[Course('SPAN', 'literature1')].prereqs
        literature_count = 3
        lingusitc_req = self.span_dict[Course('SPAN', 'linguistic')].prereqs
        lingusitc_count = 1
        electives_req = self.span_dict[Course('SPAN', 'electives1')].prereqs
        electives_count = 3
        courses = []
        for scheduled in plan:
            course = scheduled[0]
            courses.append(course)
            full_name = course[0] + course[1]
            if full_name in literature_req:
                literature_count -= 1
            if full_name in lingusitc_req:
                electives_count -= 1
            if full_name in electives_req:
                electives_count -= 1
        # check that all non-core requirements are satisfied
        self.assertTrue(literature_count <= 0)
        self.assertTrue(lingusitc_count <= 0)
        self.assertTrue(electives_count <= 0)
        # check that there are no duplicates
        self.assertNotEqual(len(courses), len(set(courses)))
    """

def split_by_term(plan):
    """
    a function that is helpful for many tests
    given a plan, returns a dictionary of scheduled courses organized by year and term
    """
    scheduled_by_term = {'Frosh': {'Fall': {'courses': [], 'credits': 0},'Spring': {'courses': [], 'credits': 0}},
                         'Soph': {'Fall': {'courses': [], 'credits': 0}, 'Spring': {'courses': [], 'credits': 0}},
                         'Junior': {'Fall': {'courses': [], 'credits': 0}, 'Spring': {'courses': [], 'credits': 0}},
                         'Senior': {'Fall': {'courses': [], 'credits': 0}, 'Spring': {'courses': [], 'credits': 0}}}

    for scheduled in plan:
        year = scheduled[1][1]
        term = scheduled[1][0]
        course = scheduled[0]
        credit = scheduled[2]
        scheduled_by_term[year][term]['courses'].append(course)
        scheduled_by_term[year][term]['credits'] += credit

    return scheduled_by_term

def add_span_major():
    """
    returns the course dictionary with the spanish major added
    """
    span_dict = course_dictionary.create_course_dict()
    Course = namedtuple('Course', 'program, designation')
    CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')

    # Core requirements: 3301W, 3302, and 3303
    core_course = Course('SPAN', 'core')
    core_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','3301W'), Course('SPAN','3302'), Course('SPAN','3303')))
    span_dict[core_course] = core_info

    # Literature: 9 credit hours from courses numbered 4400–4980 or 3835, or 3893
    literature_course = Course('SPAN', 'literature')
    literature_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','literature1'), Course('SPAN','literature2'), Course('SPAN','literature3')))
    span_dict[literature_course] = literature_info

    lit1_course = Course('SPAN', 'literature1')
    lit1_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','4405'), Course('SPAN','4420') ,Course('SPAN','4425'), Course('SPAN','4620'), Course('SPAN','4670'), Course('SPAN','4750'), Course('SPAN','4760'), Course('SPAN','4415'), Course('SPAN','4450'), Course('SPAN','4465'), Course('SPAN','4475'), Course('SPAN','4550'), Course('SPAN','4740'), Course('SPAN','4355'), Course('SPAN','4640'), Course('SPAN','4730'), Course('SPAN','4741'), Course('SPAN','4400'), Course('SPAN','4440')))
    span_dict[lit1_course] = lit1_info

    lit2_course = Course('SPAN', 'literature2')
    lit2_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','4405'), Course('SPAN','4420') ,Course('SPAN','4425'), Course('SPAN','4620'), Course('SPAN','4670'), Course('SPAN','4750'), Course('SPAN','4760'), Course('SPAN','4415'), Course('SPAN','4450'), Course('SPAN','4465'), Course('SPAN','4475'), Course('SPAN','4550'), Course('SPAN','4740'), Course('SPAN','4355'), Course('SPAN','4640'), Course('SPAN','4730'), Course('SPAN','4741'), Course('SPAN','4400'), Course('SPAN','4440')))
    span_dict[lit2_course] = lit2_info

    lit3_course = Course('SPAN', 'literature3')
    lit3_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','4405'), Course('SPAN','4420') ,Course('SPAN','4425'), Course('SPAN','4620'), Course('SPAN','4670'), Course('SPAN','4750'), Course('SPAN','4760'), Course('SPAN','4415'), Course('SPAN','4450'), Course('SPAN','4465'), Course('SPAN','4475'), Course('SPAN','4550'), Course('SPAN','4740'), Course('SPAN','4355'), Course('SPAN','4640'), Course('SPAN','4730'), Course('SPAN','4741'), Course('SPAN','4400'), Course('SPAN','4440')))
    span_dict[lit3_course] = lit3_info

    # Linguistics: 3 credit hours from courses numbered 4300–4360, or 3892
    linguistic_course = Course('SPAN', 'linguistic')
    linguistic_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','3892'),Course('SPAN','4315'),Course('SPAN','4320'),Course('SPAN','4340'),Course('SPAN','4345'),Course('SPAN','4310'),Course('SPAN','4335'),Course('SPAN','4355'),Course('SPAN','4325')))
    span_dict[linguistic_course] = linguistic_info

    # Electives: 9 credit hours from courses numbered 3320–3835 or 3891-4980
    elective_course = Course('SPAN', 'electives')
    elective_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','electives1'), Course('SPAN','electives2'), Course('SPAN','electives3')))
    span_dict[elective_course] = elective_info

    elective1_course = Course('SPAN', 'electives1')
    elective1_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','3340'),Course('SPAN','3345'),Course('SPAN','3355'),Course('SPAN','3360'),Course('SPAN','3830'),Course('SPAN','3365'),Course('SPAN','3835'),Course('SPAN','3325'),Course('SPAN','3330'),Course('SPAN','3350')))
    span_dict[elective1_course] = elective1_info

    elective2_course = Course('SPAN', 'electives2')
    elective2_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','3340'),Course('SPAN','3345'),Course('SPAN','3355'),Course('SPAN','3360'),Course('SPAN','3830'),Course('SPAN','3365'),Course('SPAN','3835'),Course('SPAN','3325'),Course('SPAN','3330'),Course('SPAN','3350')))
    span_dict[elective2_course] = elective2_info

    elective3_course = Course('SPAN', 'electives3')
    elective3_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','3340'),Course('SPAN','3345'),Course('SPAN','3355'),Course('SPAN','3360'),Course('SPAN','3830'),Course('SPAN','3365'),Course('SPAN','3835'),Course('SPAN','3325'),Course('SPAN','3330'),Course('SPAN','3350')))
    span_dict[elective3_course] = elective3_info

    major_course = Course('SPAN', 'major')
    major_info = CourseInfo(0, 'Spring, Fall', (Course('SPAN','core'),Course('SPAN','literature'), Course('SPAN','linguistic'), Course('SPAN','electives')))
    span_dict[major_course] = major_info

    return span_dict

class Group10Tests(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # create dictionaries using all the different catalogs Group 10 made
        cls.catalog1 = create_course_dict('group10tests/1.xlsx')
        cls.catalog2 = create_course_dict('group10tests/2.xlsx')
        cls.catalog3 = create_course_dict('group10tests/3.xlsx')
        cls.catalog4 = create_course_dict('group10tests/4.xlsx')
        cls.catalog5 = create_course_dict('group10tests/5.xlsx')
        cls.catalog6 = create_course_dict('group10tests/6.xlsx')
        cls.catalog7 = create_course_dict('group10tests/7.xlsx')
        cls.catalog8 = create_course_dict('group10tests/8.xlsx')
        cls.catalog9 = create_course_dict('group10tests/9.xlsx')
        cls.catalog10 = create_course_dict('group10tests/10.xlsx')
        cls.catalog11 = create_course_dict('group10tests/11.xlsx')

    """
    def test_all(self):
        plan1 = course_scheduler(self.catalog1, [('CS', 'major')], [])
        plan2 = course_scheduler(self.catalog2, [('CS', 'major')], [])
        plan3 = course_scheduler(self.catalog3, [('CS', 'major')], [])
        plan4 = course_scheduler(self.catalog4, [('CS', 'major')], [])
        plan5 = course_scheduler(self.catalog5, [('CS', 'major')], [])
        plan6 = course_scheduler(self.catalog6, [('CS', 'major')], [])
        plan7 = course_scheduler(self.catalog7, [('CS', 'major')], [])
        plan8 = course_scheduler(self.catalog8, [('CS', 'major')], [])
        plan9 = course_scheduler(self.catalog9, [('CS', 'major')], [])
        plan10 = course_scheduler(self.catalog10, [('CS', 'major')], [])
        plan11 = course_scheduler(self.catalog11, [('CS', 'major')], [])
    """


def create_course_dict(fname):
    """
    (this is copied from the provided readcsv.py, but slightly modified so that any course catalog file can be used)

    Creates a dictionary containing course info.
    Keys: namedtuple of the form ('program, designation')
    Values: namedtuple of the form('name, prereqs, credits')
            prereqs is a tuple of prereqs where each prereq has the same form as the keys
    """
    warnings.simplefilter("ignore")
    wb = load_workbook(fname)
    catalog = wb.get_sheet_by_name('catalog')
    Course = namedtuple('Course', 'program, designation')
    CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
    course_dict = {}
    for row in range(1, catalog.max_row + 1):
        key = Course(course_dictionary.get_val(catalog, 'A', row), course_dictionary.get_val(catalog, 'B', row))
        prereqs = tuple(tuple(course_dictionary.get_split_course(prereq) for prereq in prereqs.split())
                        for prereqs in course_dictionary.none_split(course_dictionary.get_val(catalog, 'E', row)))
        val = CourseInfo(course_dictionary.get_val(catalog, 'C', row), tuple(course_dictionary.get_val(catalog, 'D', row).split()), prereqs)
        course_dict[key] = val
    return course_dict


if __name__ == "__main__":
    unittest.main()
